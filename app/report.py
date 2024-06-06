import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from models import Order


def generate_report():
    orders = Order.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Order ID", "Order Name", "Description", "Creation Date", "Status"])

    status_colors = {
        "New": "0000FF",  # Blue
        "In Progress": "FFFF00",  # Yellow
        "Completed": "00FF00",  # Green
    }

    for order in orders:
        row = [
            order.id,
            order.name,
            order.description,
            order.creation_date,
            order.status,
        ]
        sheet.append(row)

        # Apply color to the row based on status
        if order.status in status_colors:
            fill_color = PatternFill(
                start_color=status_colors[order.status],
                end_color=status_colors[order.status],
                fill_type="solid",
            )
            for cell in sheet[sheet.max_row]:
                cell.fill = fill_color

    report_path = os.path.join(os.getcwd(), "order_report.xlsx")
    workbook.save(report_path)
    return report_path
