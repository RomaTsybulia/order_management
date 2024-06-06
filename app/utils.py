import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from models import Order
from database import db


def generate_report():
    orders = Order.query.all()
    wb = Workbook()
    ws = wb.active
    ws.append(["Order ID", "Order Name", "Description", "Creation Date", "Status"])

    status_colors = {"New": "0000FF", "In Progress": "FFFF00", "Completed": "00FF00"}

    for order in orders:
        row = [
            order.id,
            order.name,
            order.description,
            order.creation_date.strftime("%Y-%m-%d"),
            order.status,
        ]
        ws.append(row)
        for cell in ws[-1]:
            if cell.value == order.status:
                cell.fill = PatternFill(
                    start_color=status_colors[order.status],
                    end_color=status_colors[order.status],
                    fill_type="solid",
                )

    file_path = "../order_report.xlsx"
    wb.save(file_path)
    return file_path


def get_order_statistics():
    orders = Order.query.all()
    df = pd.DataFrame([order.serialize() for order in orders])
    stats = df["status"].value_counts().to_dict()
    return stats


def export_to_hdf5():
    orders = Order.query.all()
    df = pd.DataFrame([order.serialize() for order in orders])
    file_path = "../orders.h5"
    df.to_hdf(file_path, key="orders", mode="w")
    return file_path


def import_from_hdf5(file_path):
    df = pd.read_hdf(file_path, "orders")
    for _, row in df.iterrows():
        order = Order(
            name=row["name"], description=row["description"], status=row["status"]
        )
        db.session.add(order)
    db.session.commit()


def export_to_xml():
    orders = Order.query.all()
    import xml.etree.ElementTree as ET

    root = ET.Element("Orders")
    for order in orders:
        order_elem = ET.SubElement(root, "Order")
        ET.SubElement(order_elem, "OrderID").text = str(order.id)
        ET.SubElement(order_elem, "OrderName").text = order.name
        ET.SubElement(order_elem, "Description").text = order.description
        ET.SubElement(order_elem, "CreationDate").text = order.creation_date.strftime(
            "%Y-%m-%d"
        )
        ET.SubElement(order_elem, "Status").text = order.status

    tree = ET.ElementTree(root)
    file_path = "../orders.xml"
    tree.write(file_path)
    return file_path


def import_from_xml(file_path):
    import xml.etree.ElementTree as ET

    tree = ET.parse(file_path)
    root = tree.getroot()

    for order_elem in root.findall("Order"):
        order = Order(
            name=order_elem.find("OrderName").text,
            description=order_elem.find("Description").text,
            status=order_elem.find("Status").text,
        )
        db.session.add(order)
    db.session.commit()
